#imports
from selenium import webdriver
import openpyxl
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

#   important ---> xlsx sheet must container 2 columns
#   1st row 1st column contain heaeding Bond Numbers
#   1st row 2nd column contain Bond Status
#   2nd (row till end) 1st column contain 123456 (6-digit bond numbers)**,
#   2nd (row and row++) 2nd column is empty , it will automatically filled after the code execution

path = "./sheets/a.xlsx"                #path of xl sheet (sheet which contain bond number)
my_wb = openpyxl.load_workbook(path)
my_sheet = my_wb.active
my_row = my_sheet.max_row           #excel sheet row count

bondNumber = []
bondStatus = []

for i in range(2, my_row + 1):                              #adding excel sheet bond numbers into list
   cell_obj = my_sheet.cell(row = i, column = 1)
   bondNumber.append(cell_obj.value)

driver = webdriver.Chrome(r'C:\Users\Saad\Desktop\chromedriver_win32\chromedriver.exe')     #Download Chromedriver and add path here

driver.get('https://prizebond.vrptechnologies.com/Search/Simple')           #website url where to check result


select = Select(driver.find_element_by_id('cphContent_ddlBond'))
select.select_by_visible_text('750')         #Bond Amount e.g Rs 200, Rs 750 , Rs 1500

for i in bondNumber:
    #Send keyboard input to the search box
    driver.find_element_by_id('cphContent_txtNumber').send_keys(i)
    driver.find_element_by_id('cphContent_btnSubmit').click()
    elements = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "message")))
    text = driver.find_element_by_id("message")
    bondStatus.append(text.text)
    blank = driver.find_element_by_id('cphContent_txtNumber')
    blank.clear()

leng = len(bondNumber)

for i in range(0,leng):
    val = my_sheet.cell(row = i+2 ,column=2)
    # print(val)

    val.value = bondStatus[i]
    # print(val.value)

    my_wb.save("./sheets/a.xlsx")           #staus will be saved in xlsx file with respect to bond numbers